import xlrd
import openpyxl
import numpy as np

from pathlib import Path
from collections import defaultdict

# # def getAroi():
# workbook = xlrd.open_workbook('curvefit_forAroi.xlsx')
# worksheet = workbook.sheet_by_name('Sheet1')
# num_rows = worksheet.nrows - 1
# curr_row = -1
# value = worksheet.cell(3,1)
# print(value)
#   # return

# aroi_dict = {}
# file = Path('curvefit_forAroi.xlsx')
# workbook = openpyxl.load_workbook(file)
# worksheet = workbook.active
# # print(worksheet)
# # print(worksheet["C4"].value)

# for row in worksheet.iter_rows(max_row=10):
#   for cell in row:
#       print(cell.value, end=" ")
#   print()

def getAroi():
  aroi = {1.5: {'0':'0.001662', '10':'0.00181', '20':'0.001963', '30':'0.001963', '40':'0.001963', '50':'0.002124', '60':'0.00229', '70':'0.002376'},
          2.0: {'0':'0.001662', '10':'0.00181', '20':'0.001963', '30':'0.001963', '40':'0.002124', '50':'0.002376', '60':'0.002463', '70':'0.002642'},
          2.5: {'0':'0.00181', '10':'0.00229', '20':'0.002463', '30':'0.002463', '40':'0.00229', '50':'0.002463', '60':'0.002642', '70':'0.002734'},
          3.0: {'0':'0.002124', '10':'0.002463', '20':'0.002827', '30':'0.002827', '40':'0.002376', '50':'0.002376', '60':'0.002642', '70':'0.002734'},}
  return aroi

# g = getAroi()
# print(g[1.5]['40'])
# print(g[2.5]['60'])